import os
import sys
print("=" * 70)
print("CERTIFICATE GENERATOR - TEST SCRIPT")
print("=" * 70)
print("\n[Test 1] Checking Python version...")
python_version = sys.version_info
print(f"Python {python_version.major}.{python_version.minor}.{python_version.micro}")
if python_version.major >= 3 and python_version.minor >= 6:
    print("Python version is compatible")
else:
    print("Python version too old. Need Python 3.6+")
    sys.exit(1)

print("\n[Test 2] Checking required files...")
required_files = [
    "certificates.py",
    "certificates.png",
    "Lora-Bold.ttf",
    "sample certificates.xlsx"
]

all_files_present = True
for file in required_files:
    if os.path.exists(file):
        size = os.path.getsize(file)
        print(f"{file} ({size:,} bytes)")
    else:
        print(f"{file} - NOT FOUND!")
        all_files_present = False

if not all_files_present:
    print("\nSome required files are missing!")
    print("Please ensure all files are in the same directory.")
else:
    print("\nAll required files present")

print("\n[Test 3] Checking Python packages...")
packages_to_test = [
    ("PIL", "Pillow"),
    ("openpyxl", "openpyxl"),
    ("img2pdf", "img2pdf"),
    ("numpy", "numpy")
]

missing_packages = []
for module_name, package_name in packages_to_test:
    try:
        __import__(module_name)
        print(f"{package_name} installed")
    except ImportError:
        print(f"{package_name} not installed (will be auto-installed)")
        missing_packages.append(package_name)

if missing_packages:
    print(f"\nThe following packages will be auto-installed when you run certificates.py:")
    for pkg in missing_packages:
        print(f"- {pkg}")


print("\n[Test 4] Analyzing certificate template...")
try:
    from PIL import Image
    cert = Image.open("certificates.png")
    width, height = cert.size
    mode = cert.mode
    print(f"Template loaded successfully")
    print(f"Dimensions: {width}px x {height}px")
    print(f"Color mode: {mode}")
    
   
    if width < 1000 or height < 700:
        print(f"Template is quite small. Consider using higher resolution.")
    else:
        print(f"Template resolution is good")
    
    cert.close()
except Exception as e:
    print(f"Error loading template: {e}")


print("\n[Test 5] Checking font file...")
try:
    from PIL import ImageFont
    font = ImageFont.truetype("Lora-Bold.ttf", 50)
    print(f"Font file loaded successfully")
except Exception as e:
    print(f"Error loading font: {e}")


print("\n[Test 6] Checking Excel file...")
try:
    from openpyxl import load_workbook
    wb = load_workbook("sample certificates.xlsx", data_only=True)
    sheet_names = wb.sheetnames
    print(f"Excel file loaded successfully")
    print(f"Sheets found: {', '.join(sheet_names)}")
    
    ws = wb[sheet_names[0]]
    row_count = 0
    for r in range(3, 100):
        if ws.cell(row=r, column=1).value is None:
            break
        row_count += 1
    
    print(f"Found {row_count} entries to process")

    if row_count > 0:
        print(f"\nSample entries:")
        for r in range(3, min(6, 3 + row_count)):
            member_id = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            print(f"    Row {r}: {member_id} - {name}")
    
    wb.close()
except Exception as e:
    print(f"  [Error] Failed loading Excel file: {e}")

print("\n[Test 7] Checking configuration files...")
config_files = [
    "CONFIGURATION_GUIDE.md",
    "QUICK_SETUP.md",
    "CHANGES_SUMMARY.md",
    "VISUAL_GUIDE.txt",
    "text_placement_config.example.json"
]

for file in config_files:
    if os.path.exists(file):
        print(f"{file}")
    else:
        print(f"{file} - Not found (optional)")


print("\n[Test 8] Checking runtime configuration...")
if os.path.exists("text_placement_config.json"):
    print(f"text_placement_config.json exists")
    print(f"Configuration is already customized")
else:
    print(f"text_placement_config.json not found")
    print(f"Will be auto-created on first run with default settings")


print("\n" + "=" * 70)
print("TEST SUMMARY")
print("=" * 70)

if all_files_present:
    print("\nAll critical files are present")
    print("System is ready to generate certificates!")
    print("\nNext steps:")
    print("1. Review QUICK_SETUP.md for your certificate layout")
    print("2. Run: python certificates.py")
    print("3. Select your Excel file")
    print("4. Check the generated certificates")
    print("5. Adjust text_placement_config.json if needed")
else:
    print("\nSome files are missing")
    print("Please ensure all required files are present before running")

print("\n" + "=" * 70)
print("Happy Certificate Generating!")
print("=" * 70)
