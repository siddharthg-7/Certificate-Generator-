import os
import sys
import re
import json
import random
import subprocess
import tkinter as tk

def ensure_dependencies():
    required = {"Pillow": "PIL", "openpyxl": "openpyxl", "img2pdf": "img2pdf"}
    missing = []
    for pkg, mod in required.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"Installing missing required packages: {', '.join(missing)}...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
            print("Dependencies installed successfully.\n")
        except Exception as e:
            sys.exit(f"Failed to automatically install dependencies: {e}\nPlease run: pip install {' '.join(missing)}")

ensure_dependencies()

from PIL import Image, ImageDraw, ImageFont
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from img2pdf import convert

TEMPLATE_FILENAME = "certificates.png"
FONT_FILENAME = "Lora-Bold.ttf"
START_ROW = 3

class EfficientTextPlacer:
    def __init__(self, config_dir=None):
        self.config_dir = config_dir or os.getcwd()
        self.config_file = os.path.join(self.config_dir, "text_placement_config.json")
        self.placement_config = self.load_or_create_config()
    
    def load_or_create_config(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            default_config = {
                "text_zones": [
                    {
                        "name": "center_middle",
                        "x_range": [0.25, 0.75],
                        "y_range": [0.45, 0.60],
                        "priority": 1,
                        "description": "Standard center placement for most certificates"
                    }
                ],
                "font_size_range": [60, 90],
                "humanize_offset": True,
                "offset_range": [-5, 5]
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, indent=4)
            return default_config

    def get_adaptive_font_size(self, text, font_path):
        f_range = self.placement_config.get("font_size_range", [60, 90])
        min_s, max_s = f_range
        length = len(text)
        if length <= 10: size = max_s
        elif length <= 15: size = int(max_s - (max_s - min_s) * 0.3)
        elif length <= 20: size = int(max_s - (max_s - min_s) * 0.6)
        else: size = min_s
        return ImageFont.truetype(font_path, size)

    def get_text_coordinates(self, text_width, text_height, image_width, image_height, bbox_left=0, bbox_top=0):
        zones = self.placement_config.get("text_zones", [])
        if zones:
            # Select zone with highest priority (lowest numerical priority value)
            active_zone = sorted(zones, key=lambda z: z.get("priority", 99))[0]
            x_range = active_zone.get("x_range", [0.25, 0.75])
            y_range = active_zone.get("y_range", [0.45, 0.60])
            
            # Calculate midpoint of the active zone in relative coordinates
            x_mid = (x_range[0] + x_range[1]) / 2.0
            y_mid = (y_range[0] + y_range[1]) / 2.0
            
            x = x_mid * image_width - text_width / 2.0
            y = y_mid * image_height - text_height / 2.0
        else:
            # Default to absolute center of template
            x = (image_width - text_width) / 2.0
            y = (image_height - text_height) / 2.0

        # Adjust for top-left glyph bounding box offsets
        x -= bbox_left
        y -= bbox_top

        # Apply random organic variations if enabled
        if self.placement_config.get("humanize_offset", False):
            min_off, max_off = self.placement_config.get("offset_range", [-5, 5])
            x += random.randint(min_off, max_off)
            y += random.randint(min_off, max_off)

        return x, y

def generate_certificate(name, template, placer, font_path, img_out, pdf_out):
    font = placer.get_adaptive_font_size(name, font_path)
    width, height = template.size
    
    temp_draw = ImageDraw.Draw(template)
    bbox = temp_draw.textbbox((0, 0), name, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    x, y = placer.get_text_coordinates(
        text_width, text_height, width, height, bbox_left=bbox[0], bbox_top=bbox[1]
    )

    # Copy base template and draw text directly for performance
    cert_img = template.copy()
    draw = ImageDraw.Draw(cert_img)
    draw.text((x, y), name, fill=(40, 40, 40, 255), font=font)

    final = cert_img.convert('RGB')
    final.save(img_out)
    
    with open(pdf_out, "wb") as f:
        f.write(convert(img_out))
        
    final.close()
    cert_img.close()

def sanitize_filename(filename):
    # Replace characters that are illegal in Windows file paths
    return re.sub(r'[\\/*?:"<>|]', "_", filename)

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide default root window to prevent background blank window
    file_path = askopenfilename(title='Select Excel Workbook', filetypes=[('Excel', '*.xlsx')])
    if not file_path: sys.exit("No file selected.")
    
    dir_path = os.path.dirname(os.path.abspath(file_path))
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    t_path = os.path.join(script_dir, TEMPLATE_FILENAME)
    f_path = os.path.join(script_dir, FONT_FILENAME)
    
    if not os.path.exists(t_path): sys.exit(f"Template not found: {t_path}")
    if not os.path.exists(f_path): sys.exit(f"Font file not found: {f_path}")
    
    paths = {"img": os.path.join(dir_path, "Output", "Images"), "pdf": os.path.join(dir_path, "Output", "PDFs")}
    for p in paths.values(): os.makedirs(p, exist_ok=True)
    
    base_template = Image.open(t_path).convert('RGBA')
    placer = EfficientTextPlacer(config_dir=script_dir)
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    count = 0
    try:
        for r in range(START_ROW, ws.max_row + 1):
            member_id = str(ws.cell(row=r, column=1).value or "").strip()
            name = str(ws.cell(row=r, column=2).value or "").strip().title()
            if not name: continue
            
            raw_out_name = member_id if member_id else name.replace(" ", "_")
            out_name = sanitize_filename(raw_out_name)
            
            img_target = os.path.join(paths["img"], f"{out_name}.png")
            pdf_target = os.path.join(paths["pdf"], f"{out_name}.pdf")
            
            try:
                generate_certificate(name, base_template, placer, f_path, img_target, pdf_target)
                print(f"Generated: {name} -> {out_name}")
                count += 1
            except Exception as e:
                print(f"Error processing {name}: {e}")
    finally:
        wb.close()
        base_template.close()

    print(f"\nSuccessfully generated {count} certificates.")
